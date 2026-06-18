#!/usr/bin/env python3
"""Extract all unique text values from an xlsx workbook to texts.json.

Usage:
    python extract.py <input.xlsx>

Outputs texts.json in the current working directory:
    {"texts": ["value1", "value2", ...]}
"""
import json
import sys

from openpyxl import load_workbook


def extract_unique_texts(path: str) -> list[str]:
    wb = load_workbook(path, data_only=False)
    texts: set[str] = set()

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    value = cell.value.strip()
                    if value:
                        texts.add(value)

    return sorted(texts)


def main():
    if len(sys.argv) != 2:
        print("Usage: python extract.py <input.xlsx>", file=sys.stderr)
        sys.exit(1)

    input_path = sys.argv[1]
    texts = extract_unique_texts(input_path)

    with open("texts.json", "w", encoding="utf-8") as f:
        json.dump({"texts": texts}, f, ensure_ascii=False, indent=2)

    print(f"Extracted {len(texts)} unique texts to texts.json")


if __name__ == "__main__":
    main()
