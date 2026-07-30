
from email.mime import text
import html
import json
import os
import re
import zipfile
import time
import dotenv
from openai import OpenAI
from openpyxl import load_workbook

dotenv.load_dotenv()

def translate_excel(path: str):
  """
    Extract all unique text values from an Excel file.
  """
  wb = load_workbook(path)
  texts = set()

  for sheet in wb.worksheets:
      # 添加 sheet 名称到 texts 集合中
      if sheet.title:
        texts.add(sheet.title)

      for row in sheet.iter_rows():
          for cell in row:
              if isinstance(cell.value, str):
                  value = cell.value.strip()

                  if value:
                      texts.add(value)

  return texts

def translate_word(path: str):
  """
    Extract all unique text values from a .docx at the zip/XML level.
    按 <w:t> run 边界提取，替换时同样以 <w:t> 为单位，从而完整保留
    段落、样式、图片、域等结构（不做 run 合并，避免破坏格式）。
  """
  texts = set()
  with zipfile.ZipFile(path, "r") as z:
      for info in z.infolist():
          name = info.filename
          if not name.startswith("word/") or not name.endswith(".xml"):
              continue
          base = name[len("word/"):]
          # 只处理承载正文/页眉/页脚/脚注/尾注文本的部件，避免改动设置等部件
          if not (base == "document.xml"
                  or base.startswith(("header", "footer", "footnotes", "endnotes"))):
              continue
          xml_text = z.read(name).decode("utf-8", errors="ignore")
          for m in _W_T_PATTERN.finditer(xml_text):
              norm = _normalize(html.unescape(m.group(2)))
              if norm:
                  texts.add(norm)
  return texts

def translate_ppt(path: str):
  """
    Extract all unique text values from a .pptx at the zip/XML level.
    按 <a:t> run 边界提取（DrawingML 文本运行），替换时同样以 <a:t> 为单位，
    完整保留版式、母版、动画、图表等结构（不做 run 合并）。
    只处理正文幻灯片与备注页，不改动母版/版式（其文字多为占位提示，且被多页共用）。
  """
  texts = set()
  with zipfile.ZipFile(path, "r") as z:
      for info in z.infolist():
          name = info.filename
          if not (name.startswith("ppt/slides/") or name.startswith("ppt/notesSlides/")) \
                  or not name.endswith(".xml"):
              continue
          xml_text = z.read(name).decode("utf-8", errors="ignore")
          for m in _A_T_PATTERN.finditer(xml_text):
              norm = _normalize(html.unescape(m.group(2)))
              if norm:
                  texts.add(norm)
  return texts

# separator = '======'

def write_translate_json(json_path: str, texts: set):
    """把唯一文本集合写成 {原文: ""} 的 JSON，供大模型填充 value。"""
    ordered = sorted(texts)
    with open(json_path, "w", encoding="utf-8") as f:
        f.write('{')
        for i, text in enumerate(ordered):
            f.write(f'{json.dumps(text)}: ""')
            if i < len(ordered) - 1:
                f.write(',')
        f.write('}')

def print_translate(json_path: str, path: str):
    write_translate_json(json_path, translate_excel(path))

def translator(json_file: str, src: str, dest: str):
    """
        str: zh-cn
        dist: en
    """
    from googletrans import Translator  # 惰性导入：避免与 openai 的 httpx 版本冲突
    translator = Translator()

    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    translated = {}
    total = len(data)

    for i, (key, value) in enumerate(data.items(), 1):
        if value:  # Skip if already translated
            translated[key] = value
            continue

        try:
            result = translator.translate(key, src, dest)
            translated[key] = result.text
            print(f"[{i}/{total}] Translated: {key[:50]}... -> {result.text[:50]}...")
        except Exception as e:
            print(f"[{i}/{total}] Error translating: {key[:50]}... - {e}")
            translated[key] = key  # Keep original on error

        # Small delay to avoid rate limiting
        time.sleep(0.3)

    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(translated, f, ensure_ascii=False, indent=2)

    print(f"\nDone! Translated {total} entries.")

_T_PATTERN = re.compile(r'(<t(?:\s[^>]*)?>)([^<>]*)(</t>)')
_SHEET_NAME_PATTERN = re.compile(r'(<sheet\b[^>]*?\sname=")([^"]*)(")')
# Word 正文文本运行：<w:t>...</w:t>（含 <w:t xml:space="preserve">...</w:t>）
_W_T_PATTERN = re.compile(r'(<w:t(?:\s[^>]*)?>)([^<>]*)(</w:t>)')
# PPT 正文文本运行：<a:t>...</a:t>（DrawingML，含 <a:t xml:space="preserve">...</a:t>）
_A_T_PATTERN = re.compile(r'(<a:t(?:\s[^>]*)?>)([^<>]*)(</a:t>)')
# 匹配 JSON 对象里「key 缺少 : value」的残缺条目：键后直接是 , 或 }
#   ..."key",  → ..."key": "key",
# 只匹配键位置（前驱是 { 或 ,），不会误伤值（值前驱是 :）。
_KEY_MISSING_VALUE = re.compile(r'([{,]\s*)("(?:[^"\\]|\\.)*")(\s*)([,}])')


def _normalize(content: str) -> str:
    return content.replace('\r\n', '\n').replace('\r', '\n').strip()


def _parse_json(raw: str):
    """从大模型原始输出解析 JSON，容错 markdown 围栏与「key 缺值」残缺条目。"""
    s = raw.strip()
    if s.startswith("```"):
        lines = s.splitlines()
        if lines and lines[0].lstrip().startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].lstrip().startswith("```"):
            lines = lines[:-1]
        s = "\n".join(lines).strip()
    start, end = s.find("{"), s.rfind("}")
    if start != -1 and end > start:
        s = s[start:end + 1]
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        return json.loads(_KEY_MISSING_VALUE.sub(r'\1\2: \2\4', s))


def _load_translated_map(raw: str, original_keys) -> dict:
    """解析 LLM 输出并与原始 key 对齐：缺失或空值用 key 自身兜底。"""
    data = _parse_json(raw)
    result = {}
    for k in original_keys:
        v = data.get(k) if isinstance(data, dict) else None
        result[k] = v if v else k
    return result


# 匹配 CJK 汉字（中文源方向：value 仍含汉字 = 没翻出来）
_CJK = re.compile(r'[一-鿿㐀-䶿]')
# 匹配拉丁字母（英文源方向：原文与译文一致且含字母 = LLM 原样回吐，没翻）
_LATIN = re.compile(r'[A-Za-z]')


def _needs_translation(key: str, value: str) -> bool:
    """value 是否仍需翻译。

    中文源方向：value 仍含汉字 → 没翻。
    英文源方向：value 与 key 完全一致且含拉丁字母 → LLM 原样回吐，没翻。
    纯数字/型号编号/单位符号（无汉字、无字母，或与 key 一致但本就允许保留）不会被命中。
    """
    if _CJK.search(value):
        return True
    if value == key and _LATIN.search(value):
        return True
    return False


def _build_translate_prompt(translate: str, data: dict) -> str:
    return f"""
                待翻译的 JSON 数据如下：
                ```json
                {json.dumps(data)}
                ```
                将上面 json 中的每个 key 由{translate}，翻译结果写入该 key 对应的 value 中。
                翻译规则：
                - 凡是含有源语言文字的 key，都必须翻译，不得保留原文。即便是短句、含错别字、含标点、含型号夹杂的文字，也要翻译其中的人类语言部分。
                - 只有纯数字、纯型号编号、纯单位符号（如 16.2V、3680mAh、>85%、/）才允许 value 等于 key 原样保留。
                - value 不得为空字符串。
                - key 必须原样保留，不得改写、不得纠正错别字、不得增删任何字符（否则会无法对回原文）。
                严格要求：
                1. 只返回 JSON。
                2. 不允许 Markdown 代码块。
                3. 不允许任何解释文字。
                4. 保持 JSON 结构不变，key 集合与输入完全一致。
                5. 输出必须可以被 Python json.loads() 直接解析。
                """


def _build_retry_prompt(translate: str, keys: list) -> str:
    """二次翻译用位置对齐的数组，避开 LLM 改写 key 导致对不上的问题。"""
    return f"""
                下面是一个 JSON 数组，数组中每个元素都是需要翻译的文本（按此顺序）：
                ```json
                {json.dumps(keys, ensure_ascii=False)}
                ```
                请将数组中的每个元素{translate}，并按【完全相同的顺序】返回一个 JSON 对象，格式为 {{"items": ["译文1", "译文2", ...]}}。
                要求：
                - 每条都必须翻译，绝对不允许让译文等于原文（不得保留源语言文字）。
                - 只返回 JSON，不允许 Markdown 代码块和解释文字。
                - 输出必须可被 Python json.loads() 直接解析。
                """


def replace_t_text(xml_text: str, escaped_map: dict) -> str:
    def repl(m):
        open_tag, content, close_tag = m.group(1), m.group(2), m.group(3)
        normalized = _normalize(content)
        if normalized in escaped_map:
            return open_tag + escaped_map[normalized] + close_tag
        return m.group(0)
    return _T_PATTERN.sub(repl, xml_text)


def replace_sheet_names(xml_text: str, escaped_map: dict) -> str:
    def repl(m):
        prefix, content, suffix = m.group(1), m.group(2), m.group(3)
        normalized = _normalize(content)
        if normalized in escaped_map:
            return prefix + escaped_map[normalized] + suffix
        return m.group(0)
    return _SHEET_NAME_PATTERN.sub(repl, xml_text)


def replace_w_t_text(xml_text: str, escaped_map: dict) -> str:
    """替换 Word 的 <w:t> 文本，保留外层 run 的所有格式属性。"""
    def repl(m):
        open_tag, content, close_tag = m.group(1), m.group(2), m.group(3)
        normalized = _normalize(content)
        if normalized in escaped_map:
            return open_tag + escaped_map[normalized] + close_tag
        return m.group(0)
    return _W_T_PATTERN.sub(repl, xml_text)


def replace_a_t_text(xml_text: str, escaped_map: dict) -> str:
    """替换 PPT 的 <a:t> 文本，保留外层 run 的所有格式属性。"""
    def repl(m):
        open_tag, content, close_tag = m.group(1), m.group(2), m.group(3)
        normalized = _normalize(content)
        if normalized in escaped_map:
            return open_tag + escaped_map[normalized] + close_tag
        return m.group(0)
    return _A_T_PATTERN.sub(repl, xml_text)


def excel_cell_replace(translate_path: str, path: str):
    """
    Replace the text in an xlsx at the zip/XML level, preserving images
    (including wmf), formulas, styles, and rich-text run formatting.
    """

    with open(translate_path, "r", encoding="utf-8") as f:
        translate_map = json.load(f)    

    escaped_map = {html.escape(k, quote=False): html.escape(v, quote=False)
                   for k, v in translate_map.items()}

    # 输出文件放在与输入同目录，避免依赖启动时的 CWD
    file_name = os.path.join(os.path.dirname(path),
                            os.path.splitext(os.path.basename(path))[0] + "-translated.xlsx")
    print(f"Saving translated file to {file_name}")

    with zipfile.ZipFile(path, "r") as zin, \
         zipfile.ZipFile(file_name, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            is_strings = item.filename == "xl/sharedStrings.xml"
            is_sheet = (item.filename.startswith("xl/worksheets/")
                        and item.filename.endswith(".xml"))
            is_workbook = item.filename == "xl/workbook.xml"
            if is_strings or is_sheet:
                text = data.decode("utf-8")
                text = replace_t_text(text, escaped_map)
                data = text.encode("utf-8")
            elif is_workbook:
                text = data.decode("utf-8")
                text = replace_sheet_names(text, escaped_map)
                data = text.encode("utf-8")
            zout.writestr(item, data)
    return file_name

def word_text_replace(translate_path: str, path: str):
    """
    Replace the text in a .docx at the zip/XML level, preserving images,
    styles, fields, headers/footers and rich-text run formatting.
    与 excel_cell_replace 同构：只改 <w:t> 文本，其余部件原样写回。
    """
    with open(translate_path, "r", encoding="utf-8") as f:
        translate_map = json.load(f)

    escaped_map = {html.escape(k, quote=False): html.escape(v, quote=False)
                   for k, v in translate_map.items()}

    file_name = os.path.join(os.path.dirname(path),
                            os.path.splitext(os.path.basename(path))[0] + "-translated.docx")
    print(f"Saving translated file to {file_name}")

    with zipfile.ZipFile(path, "r") as zin, \
         zipfile.ZipFile(file_name, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            name = item.filename
            if name.startswith("word/") and name.endswith(".xml"):
                base = name[len("word/"):]
                if (base == "document.xml"
                        or base.startswith(("header", "footer", "footnotes", "endnotes"))):
                    text = data.decode("utf-8")
                    text = replace_w_t_text(text, escaped_map)
                    data = text.encode("utf-8")
            zout.writestr(item, data)
    return file_name

def ppt_text_replace(translate_path: str, path: str):
    """
    Replace the text in a .pptx at the zip/XML level, preserving images,
    charts, animations, layouts and rich-text run formatting.
    与 word/excel 同构：只改 <a:t> 文本，其余部件原样写回。
    """
    with open(translate_path, "r", encoding="utf-8") as f:
        translate_map = json.load(f)

    escaped_map = {html.escape(k, quote=False): html.escape(v, quote=False)
                   for k, v in translate_map.items()}

    file_name = os.path.join(os.path.dirname(path),
                            os.path.splitext(os.path.basename(path))[0] + "-translated.pptx")
    print(f"Saving translated file to {file_name}")

    with zipfile.ZipFile(path, "r") as zin, \
         zipfile.ZipFile(file_name, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            name = item.filename
            if (name.startswith("ppt/slides/") or name.startswith("ppt/notesSlides/")) \
                    and name.endswith(".xml"):
                text = data.decode("utf-8")
                text = replace_a_t_text(text, escaped_map)
                data = text.encode("utf-8")
            zout.writestr(item, data)
    return file_name

def chat(content: str, on_chunk=None):
    client = OpenAI(
            api_key=os.getenv("OPENAI_API_KEY"),
            base_url=os.getenv("BASE_URL"),
        )

    messages = [{"role": "system","content": "你是一个JSON生成助手，只输出合法JSON"},
                {"role": "user", "content": content}]
    completion = client.chat.completions.create(
        model=os.getenv("MODEL"),  # 您可以按需更换为其它深度思考模型
        messages=messages,
        response_format={
         "type": "json_object"
        },
        extra_body={"enable_thinking": True},
        stream=True
    )

    is_answering = False  # 是否进入回复阶段
    full_content = []  # 累积完整的回复内容
    print("\n" + "=" * 20 + "思考过程" + "=" * 20)
    for chunk in completion:
        if not chunk.choices:
            continue
        delta = chunk.choices[0].delta
        if hasattr(delta, "reasoning_content") and delta.reasoning_content is not None:
            if not is_answering:
                print(delta.reasoning_content, end="", flush=True)
                if on_chunk:
                    on_chunk({"type": "reasoning", "text": delta.reasoning_content})
        if hasattr(delta, "content") and delta.content:
            if not is_answering:
                print("\n" + "=" * 20 + "完整回复" + "=" * 20)
                is_answering = True
            print(delta.content, end="", flush=True)
            # 累积完整的回复内容，流式结束后返回给调用方
            full_content.append(delta.content)
            if on_chunk:
                on_chunk({"type": "content", "text": delta.content})

    return "".join(full_content)


def run_pipeline(path: str, translate: str, json_path: str = "translate.json", on_event=None):
    """端到端翻译流程：提取 → 大模型翻译 → 替换。on_event 回调用于推送进度。"""
    def emit(ev):
        if on_event:
            on_event(ev)

    try:
        ext = os.path.splitext(path)[1].lower()
        if ext == ".xlsx":
            extract_fn = lambda: translate_excel(path)
            replace_fn = excel_cell_replace
        elif ext == ".docx":
            extract_fn = lambda: translate_word(path)
            replace_fn = word_text_replace
        elif ext == ".pptx":
            extract_fn = lambda: translate_ppt(path)
            replace_fn = ppt_text_replace
        else:
            raise ValueError(f"不支持的文件类型: {ext}（仅支持 .xlsx / .docx / .pptx）")

        emit({"type": "status", "stage": "extract"})
        # 1. 提取文本到 json_path 中
        write_translate_json(json_path, extract_fn())

        # 2. 大模型翻译
        with open(json_path, "r", encoding="utf-8") as f:
            translate_data = json.load(f)
        emit({"type": "status", "stage": "translate", "total": len(translate_data)})
        raw = chat(_build_translate_prompt(translate, translate_data),
                   on_chunk=lambda c: emit({"type": "chunk", **c}))
        merged = _load_translated_map(raw, translate_data)
        # 首遍翻译结果先落盘（二次翻译崩了也有首遍结果可用，且便于中途查看）
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(merged, f, ensure_ascii=False, indent=2)

        # 2.7 兜底重译：value 仍含源语言文字的条目反复重译，直至全部翻完或无进展。
        max_passes = 50  # 防止异常情况下死循环的后备上限
        for pass_no in range(1, max_passes + 1):
            retry_keys = [k for k, v in merged.items() if _needs_translation(k, v)]
            if not retry_keys:
                break
            emit({"type": "status", "stage": "retranslate",
                  "pass": pass_no, "total": len(retry_keys)})
            try:
                # 用位置对齐的数组，避免 LLM 改写 key 导致对不回原文
                raw2 = chat(_build_retry_prompt(translate, retry_keys),
                            on_chunk=lambda c: emit({"type": "chunk", **c}))
                items = _parse_json(raw2)
                if isinstance(items, dict):
                    items = items.get("items")
                if isinstance(items, list):
                    improved = 0
                    for i, k in enumerate(retry_keys):
                        if i < len(items):
                            nv = items[i]
                            if nv and not _needs_translation(k, nv):
                                merged[k] = nv
                                improved += 1
                # 重译回填后重新写入 translate.json
                with open(json_path, "w", encoding="utf-8") as f:
                    json.dump(merged, f, ensure_ascii=False, indent=2)
                # 本轮无任何改善则提前退出，避免对无法翻译的条目无限重试
                if improved == 0:
                    emit({"type": "status", "stage": "retranslate-stuck",
                          "message": f"pass {pass_no} 无改善，停止重译"})
                    break
            except Exception as e:
                # 重译失败不中断主流程，保留已有结果（已落盘）
                emit({"type": "status", "stage": "retranslate-skipped",
                      "message": str(e)[:120]})
                break

        # 3. 替换文件中的文本（所有翻译都已完成才生成）
        emit({"type": "status", "stage": "replace"})
        out_file = replace_fn(json_path, path)
        emit({"type": "done", "file": os.path.basename(out_file)})
        print("Completed")
    except Exception as e:
        emit({"type": "error", "message": str(e)})
        raise