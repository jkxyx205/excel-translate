
import html
import json
import os
import re
import zipfile
import time
from googletrans import Translator

from openpyxl import load_workbook


def translate_excel(path: str):
  """
    Extract all unique text values from an Excel file.
  """
  wb = load_workbook(path)
  texts = set()

  for sheet in wb.worksheets:
      # 添加 sheet 名称到 texts 集合中
      if sheet.title:
        texts.add(sheet.title)

      for row in sheet.iter_rows():
          for cell in row:
              if isinstance(cell.value, str):
                  value = cell.value.strip()

                  if value:
                      texts.add(value)

  return texts

# separator = '======'

def print_translate(json_path: str, path: str):
    texts = translate_excel(path)
    with open(json_path, "a") as f:
        f.write('{')
        for text in sorted(texts):
            # 生成 JSON 格式的键值对。key 和 value 都是 text，key 可能会「换行」 或者出现「双引号」
            key = f'"{text.replace('\n', '\\n').replace("\"", "\\\"")}": ""'
            f.write(key)  
            # 是否是最后一个元素
            if text != sorted(texts)[-1]:
                f.write(',')
        f.write('}')

def translator(json_file: str, src: str, dest: str):
    """
        str: zh-cn
        dist: en
    """
    translator = Translator()

    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    translated = {}
    total = len(data)

    for i, (key, value) in enumerate(data.items(), 1):
        if value:  # Skip if already translated
            translated[key] = value
            continue

        try:
            result = translator.translate(key, src, dest)
            translated[key] = result.text
            print(f"[{i}/{total}] Translated: {key[:50]}... -> {result.text[:50]}...")
        except Exception as e:
            print(f"[{i}/{total}] Error translating: {key[:50]}... - {e}")
            translated[key] = key  # Keep original on error

        # Small delay to avoid rate limiting
        time.sleep(0.3)

    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(translated, f, ensure_ascii=False, indent=2)

    print(f"\nDone! Translated {total} entries.")

# def loadData(path: str):
#     with open(path, 'r') as f:
#         content = f.read()
    
#     separator = "======"
#     parts = content.split(separator)
    
#     result = {}
#     number = 1
#     for part in parts:
#         text = part.strip()
#         if text:
#             result[number] = text
#             number += 1
    
#     return result

# def get_translate_map(source_path: str, target_path: str):
#     # Load the source map
#     source_map = loadData(source_path)
#     # Load the target map
#     target_map = loadData(target_path)
    
#     # 遍历 source_map
#     combined_map = {}
#     for key, value in source_map.items():
#         combined_map[value] = target_map.get(key, value)


#     # 获取 combined_map 的大小
#     print(f"source_map map size: {len(source_map)}")
#     print(f"target_map map size: {len(target_map)}")
#     print(f"Combined map size: {len(combined_map)}")

#     return combined_map

_T_PATTERN = re.compile(r'(<t(?:\s[^>]*)?>)([^<>]*)(</t>)')
_SHEET_NAME_PATTERN = re.compile(r'(<sheet\b[^>]*?\sname=")([^"]*)(")')


def _normalize(content: str) -> str:
    return content.replace('\r\n', '\n').replace('\r', '\n').strip()


def replace_t_text(xml_text: str, escaped_map: dict) -> str:
    def repl(m):
        open_tag, content, close_tag = m.group(1), m.group(2), m.group(3)
        normalized = _normalize(content)
        if normalized in escaped_map:
            return open_tag + escaped_map[normalized] + close_tag
        return m.group(0)
    return _T_PATTERN.sub(repl, xml_text)


def replace_sheet_names(xml_text: str, escaped_map: dict) -> str:
    def repl(m):
        prefix, content, suffix = m.group(1), m.group(2), m.group(3)
        normalized = _normalize(content)
        if normalized in escaped_map:
            return prefix + escaped_map[normalized] + suffix
        return m.group(0)
    return _SHEET_NAME_PATTERN.sub(repl, xml_text)


def excel_cell_replace(translate_path: str, path: str):
    """
    Replace the text in an xlsx at the zip/XML level, preserving images
    (including wmf), formulas, styles, and rich-text run formatting.
    """

    with open(translate_path, "r", encoding="utf-8") as f:
        translate_map = json.load(f)    

    escaped_map = {html.escape(k, quote=False): html.escape(v, quote=False)
                   for k, v in translate_map.items()}

    # 获取 path 中原文件名后再添加"-translated"作为新的文件名
    file_name = os.path.splitext(os.path.basename(path))[0] + "-translated.xlsx"
    print(f"Saving translated file to {file_name}")

    with zipfile.ZipFile(path, "r") as zin, \
         zipfile.ZipFile(file_name, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            is_strings = item.filename == "xl/sharedStrings.xml"
            is_sheet = (item.filename.startswith("xl/worksheets/")
                        and item.filename.endswith(".xml"))
            is_workbook = item.filename == "xl/workbook.xml"
            if is_strings or is_sheet:
                text = data.decode("utf-8")
                text = replace_t_text(text, escaped_map)
                data = text.encode("utf-8")
            elif is_workbook:
                text = data.decode("utf-8")
                text = replace_sheet_names(text, escaped_map)
                data = text.encode("utf-8")
            zout.writestr(item, data)

if __name__ == "__main__":
    path = "./excel/TV83有刷TV83S无刷SOP-1.xlsx"
    json_path = "translate.json"

    # 1. 获取翻译的 excel 文本到 1.txt 中
    # print_translate(json_path, path)

    # 2. 大模型翻译，claude --dangerously-skip-permissions 提示词：将 @translate.json 中的 key 由中文翻译成英文，翻译结果写入 key 对应的 value 中;
    # 2. 直接调用谷歌翻译，不稳定，经常翻译不出来
    # translator(json_path, 'zh-cn', 'en')

    # 3. 替换 Excel 中的文本
    excel_cell_replace(json_path, path)

    print('Completed')
    # 注意：图片中的文字无法翻译